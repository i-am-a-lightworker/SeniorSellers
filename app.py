import io
import re
import time
import urllib.parse
import urllib.robotparser
from dataclasses import dataclass
from typing import Iterable

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

st.set_page_config(page_title="Property Sale Signal Researcher", layout="wide")

SALE_TERMS = {
    "for sale": 8,
    "coming soon": 7,
    "open house": 6,
    "fsbo": 8,
    "for sale by owner": 9,
    "listing": 3,
    "listed": 3,
    "price reduced": 5,
    "back on market": 5,
    "active listing": 6,
    "withdrawn listing": 4,
    "expired listing": 4,
    "brokerage": 2,
    "real estate": 1,
}

EXCLUDED_TERMS = {
    "foreclosure",
    "tax delinquent",
    "bankruptcy",
    "divorce",
    "probate",
    "death",
    "illness",
    "job loss",
    "financial hardship",
}

ADDRESS_CANDIDATES = ["address", "property address", "full address", "site address", "mailing address"]
STATUS_CANDIDATES = ["status", "compliance status", "cd status", "cease and desist"]
URL_CANDIDATES = ["url", "source url", "page url", "link"]


@dataclass
class FetchResult:
    url: str
    allowed: bool
    status_code: int | None
    title: str
    text: str
    error: str


def normalize_col(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip().lower())


def find_column(df: pd.DataFrame, candidates: Iterable[str]) -> str | None:
    lookup = {normalize_col(c): c for c in df.columns}
    for candidate in candidates:
        if candidate in lookup:
            return lookup[candidate]
    for normalized, original in lookup.items():
        if any(candidate in normalized for candidate in candidates):
            return original
    return None


def read_table(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return pd.read_excel(uploaded_file)
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    raise ValueError("Upload an Excel (.xlsx/.xlsm) or CSV file.")


def build_search_links(address: str) -> dict[str, str]:
    safe_address = str(address).strip()
    queries = {
        "Public sale search": f'"{safe_address}" ("for sale" OR "coming soon" OR "open house" OR FSBO)',
        "Listing search": f'"{safe_address}" real estate listing',
        "Property page search": f'"{safe_address}" property',
    }
    return {
        label: "https://www.google.com/search?q=" + urllib.parse.quote_plus(query)
        for label, query in queries.items()
    }


def robots_allows(url: str, user_agent: str = "PropertySaleSignalResearcher/1.0") -> tuple[bool, str]:
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return False, "Invalid URL"
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    rp = urllib.robotparser.RobotFileParser()
    rp.set_url(robots_url)
    try:
        rp.read()
        return rp.can_fetch(user_agent, url), ""
    except Exception as exc:
        return False, f"Could not verify robots.txt: {exc}"


def fetch_public_page(url: str, timeout: int = 12) -> FetchResult:
    allowed, robots_error = robots_allows(url)
    if not allowed:
        return FetchResult(url, False, None, "", "", robots_error or "Blocked by robots.txt")

    headers = {
        "User-Agent": "PropertySaleSignalResearcher/1.0 (+local compliance research tool)",
        "Accept": "text/html,application/xhtml+xml",
    }
    try:
        response = requests.get(url, timeout=timeout, headers=headers)
        content_type = response.headers.get("content-type", "").lower()
        if "text/html" not in content_type:
            return FetchResult(url, True, response.status_code, "", "", "Non-HTML content skipped")
        soup = BeautifulSoup(response.text, "html.parser")
        for tag in soup(["script", "style", "noscript", "svg"]):
            tag.decompose()
        title = soup.title.get_text(" ", strip=True) if soup.title else ""
        text = " ".join(soup.get_text(" ", strip=True).split())
        return FetchResult(url, True, response.status_code, title, text[:200000], "")
    except Exception as exc:
        return FetchResult(url, True, None, "", "", str(exc))


def score_text(text: str) -> tuple[int, list[str], list[str]]:
    lower = text.lower()
    found_excluded = sorted(term for term in EXCLUDED_TERMS if term in lower)
    clean_text = lower
    for term in found_excluded:
        clean_text = clean_text.replace(term, " ")

    score = 0
    found_terms = []
    for term, weight in SALE_TERMS.items():
        if term in clean_text:
            score += weight
            found_terms.append(term)
    return score, sorted(found_terms), found_excluded


def classify(score: int) -> str:
    if score >= 10:
        return "PUBLIC SALE SIGNAL"
    if score >= 4:
        return "POSSIBLE PROPERTY-LEVEL SIGNAL"
    return "NO PUBLIC SALE SIGNAL FOUND"


def export_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    df.to_excel(output, index=False, sheet_name="Results")
    output.seek(0)
    wb = load_workbook(output)
    ws = wb["Results"]
    headers = {cell.value: cell.column for cell in ws[1]}
    status_col = headers.get("FINAL STATUS")
    cd_col = headers.get("CEASE AND DESIST") or headers.get("Compliance Status")

    red_fill = PatternFill("solid", fgColor="FF0000")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    for row in range(2, ws.max_row + 1):
        cd_value = str(ws.cell(row, cd_col).value).upper() if cd_col else ""
        status_value = str(ws.cell(row, status_col).value).upper() if status_col else ""
        target = ws.cell(row, status_col) if status_col else ws.cell(row, 1)
        if "CEASE AND DESIST" in cd_value or status_value == "CEASE AND DESIST":
            target.fill = red_fill
            target.font = white_font
        elif status_value == "PUBLIC SALE SIGNAL":
            target.fill = green_fill
            target.font = bold_font
        elif "POSSIBLE" in status_value:
            target.fill = yellow_fill
            target.font = bold_font

    for column_cells in ws.columns:
        max_length = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length

    final = io.BytesIO()
    wb.save(final)
    return final.getvalue()


st.title("Property Sale Signal Researcher")
st.caption("API-free, local, property-level research. It does not profile residents or collect sensitive personal circumstances.")

st.warning(
    "This tool must not be used to infer personal hardship, health, family status, protected characteristics, "
    "or other private circumstances. A CEASE AND DESIST result always overrides every marketing signal."
)

uploaded = st.file_uploader("Upload an eligible-property list or the output from your Cease and Desist Checker", type=["xlsx", "xlsm", "csv"])

if uploaded:
    try:
        df = read_table(uploaded)
    except Exception as exc:
        st.error(str(exc))
        st.stop()

    address_col = find_column(df, ADDRESS_CANDIDATES)
    status_col = find_column(df, STATUS_CANDIDATES)
    url_col = find_column(df, URL_CANDIDATES)

    st.subheader("Column mapping")
    col1, col2, col3 = st.columns(3)
    with col1:
        address_col = st.selectbox("Address column", list(df.columns), index=list(df.columns).index(address_col) if address_col in df.columns else 0)
    with col2:
        status_options = ["None"] + list(df.columns)
        default_status = status_options.index(status_col) if status_col in df.columns else 0
        selected_status = st.selectbox("Cease-and-desist/status column", status_options, index=default_status)
        status_col = None if selected_status == "None" else selected_status
    with col3:
        url_options = ["None"] + list(df.columns)
        default_url = url_options.index(url_col) if url_col in df.columns else 0
        selected_url = st.selectbox("Optional public URL column", url_options, index=default_url)
        url_col = None if selected_url == "None" else selected_url

    max_pages = st.number_input("Maximum public pages to fetch this run", min_value=0, max_value=200, value=25, step=5)
    delay = st.slider("Delay between page requests (seconds)", 0.5, 5.0, 1.5, 0.5)

    if st.button("Run property-level signal review", type="primary"):
        results = df.copy()
        results["SEARCH LINK 1"] = ""
        results["SEARCH LINK 2"] = ""
        results["SEARCH LINK 3"] = ""
        results["SOURCE TITLE"] = ""
        results["SIGNAL SCORE"] = 0
        results["SALE TERMS FOUND"] = ""
        results["EXCLUDED SENSITIVE TERMS IGNORED"] = ""
        results["FETCH NOTE"] = ""
        results["FINAL STATUS"] = ""

        fetched = 0
        progress = st.progress(0)
        total = max(len(results), 1)

        for idx, row in results.iterrows():
            address = str(row.get(address_col, "") or "").strip()
            links = build_search_links(address)
            results.at[idx, "SEARCH LINK 1"] = links["Public sale search"]
            results.at[idx, "SEARCH LINK 2"] = links["Listing search"]
            results.at[idx, "SEARCH LINK 3"] = links["Property page search"]

            cd_value = str(row.get(status_col, "") or "").upper() if status_col else ""
            if "CEASE AND DESIST" in cd_value:
                results.at[idx, "FINAL STATUS"] = "CEASE AND DESIST"
                results.at[idx, "FETCH NOTE"] = "Blocked from marketing review"
                progress.progress((idx + 1) / total)
                continue

            combined_text = address
            source_title = ""
            fetch_note = "Manual search links generated"

            if url_col and fetched < max_pages:
                url = str(row.get(url_col, "") or "").strip()
                if url:
                    result = fetch_public_page(url)
                    fetched += 1
                    if result.text:
                        combined_text += " " + result.title + " " + result.text
                        source_title = result.title
                        fetch_note = f"Fetched HTTP {result.status_code}"
                    else:
                        fetch_note = result.error or "No text extracted"
                    time.sleep(delay)

            score, terms, excluded = score_text(combined_text)
            results.at[idx, "SOURCE TITLE"] = source_title
            results.at[idx, "SIGNAL SCORE"] = score
            results.at[idx, "SALE TERMS FOUND"] = ", ".join(terms)
            results.at[idx, "EXCLUDED SENSITIVE TERMS IGNORED"] = ", ".join(excluded)
            results.at[idx, "FETCH NOTE"] = fetch_note
            results.at[idx, "FINAL STATUS"] = classify(score)
            progress.progress((idx + 1) / total)

        st.success("Review complete")
        st.dataframe(results, use_container_width=True)

        excel_bytes = export_excel(results)
        st.download_button(
            "Download reviewed Excel file",
            data=excel_bytes,
            file_name="property_sale_signal_review.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

st.divider()
st.markdown(
    "**Permitted use:** public property listings, FSBO advertisements, coming-soon pages, open-house pages, and other property-level marketing content. "
    "**Not permitted:** social-media profiling of residents, personal-data enrichment, protected-class inference, or hardship-based targeting."
)
